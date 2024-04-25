import openpyxl
import os
import cantools
import re
import logging

def parse_dbc(dbc_file_path):
    dbc_info = {}
    signal_length_dict = {}
    # dbc_info['messages'] = []
    for file_name in os.listdir(dbc_file_path):
        if file_name.endswith('.dbc'):
            dbc_file = os.path.join(dbc_file_path, file_name)
            dbc_name = os.path.splitext(file_name)[0]
            db = cantools.database.load_file(dbc_file)

            for message in db.messages:
                message_info = {
                    'name': f'{dbc_name}_{message.name}',
                    'Cycle Time': message.cycle_time,
                    'is_fd': message.is_fd,
                    'signals' : []
                }
                dbc_info[message_info['name']] = []
                for signal in message.signals:
                    signal_info = {
                        'name': signal.name,
                        'length': signal.length,
                        'Factor': signal.scale
                    }

                    signal_length_dict[f"{message.name}_{signal.name}"] = "{:02d}".format(signal.length)

                    dbc_info[message_info['name']].append(signal.name)

    return signal_length_dict, dbc_info

# check that the messages and signals in DBC exists in excel
def excel_check(source_folder_path, excel_file):
    dbc_info = parse_dbc(source_folder_path)[1]
    workbook1 = openpyxl.load_workbook(excel_file)
    # workbook2 = openpyxl.load_workbook('./output/dbc_info.xlsx')
    new_log = open('./output/Not found.log', 'w', encoding='utf-8')
    ecu2vehicle = workbook1.worksheets[2]  
    vehicle2ecu = workbook1.worksheets[3]

    for row in ecu2vehicle.iter_rows(min_row=2, values_only=True):
        ecu_signal = row[0]
        ecu_message = f'ecu_{row[2]}'
        if (ecu_message in dbc_info) and (ecu_signal in dbc_info[ecu_message]):
            pass
        else:
            log_message_new = f"Message: {ecu_message}, Signal: {ecu_signal} is not contained in DBC."
            new_log.write(log_message_new + '\n')

        vehicle_signal = row[5]
        vehicle_message = f'vehicle_{row[7]}'
        if (vehicle_message in dbc_info) and (vehicle_signal in dbc_info[vehicle_message]):
            pass
        else:
            log_message_new = f"Message: {vehicle_message}, Signal: {vehicle_signal} is not contained in DBC."
            new_log.write(log_message_new + '\n')

    for row in vehicle2ecu.iter_rows(min_row=2, values_only=True):
        ecu_signal = row[5]
        ecu_message = f'ecu_{row[7]}'
        if (ecu_message in dbc_info) and (ecu_signal in dbc_info[ecu_message]):
            pass
        else:
            log_message_new = f"Message: {ecu_message}, Signal: {ecu_signal} is not contained in DBC."
            new_log.write(log_message_new + '\n')

        vehicle_signal = row[0]
        vehicle_message = f'vehicle_{row[2]}'
        if (vehicle_message in dbc_info) and (vehicle_signal in dbc_info[vehicle_message]):
            pass
        else:
            log_message_new = f"Message: {vehicle_message}, Signal: {vehicle_signal} is not contained in DBC."
            new_log.write(log_message_new + '\n')

    workbook1.close()
    new_log.close()

def camel_to_underscore(variable_name):
    # 将驼峰命名的变量名转换为下划线命名
    pattern = re.compile(r'(?<!^)[A-Z](?=[a-z])|(?<=[a-z])(?=[A-Z][a-z])|(?<=[a-z])(?=[A-Z])|(?<=[0-9])(?=[A-Z])')
    underscore_name = pattern.sub(lambda x: '_' + x.group().lower(), variable_name).lower()
    underscore_name = underscore_name.replace("__", "_")
    return underscore_name

class ExcelParser:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)        

    def parse_mapping_rules(self, sheet_num, signal_length_dict, device: str):
        sheet = self.workbook.worksheets[sheet_num]
        mapping_rule_column = sheet["K"]
        vehicle_message_name_column = sheet["C"]
        vehicle_signal_name_column = sheet["A"]
        ecu_message_name_column = sheet["H"]
        ecu_signal_name_column = sheet["F"]
        information = []
        infors = []
        min_rows = [0]
        max_rows = [0]
        m = 0
        merged_cells = sheet.merged_cells.ranges
        sheet.cell(1, 12).value = 'Input'
        sheet.cell(1, 13).value = 'Output'
        for merged_range in merged_cells:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_col <= 11 <= max_col:
                min_rows.append(min_row)
                max_rows.append(max_row)

        for i, cell in enumerate(mapping_rule_column[1:]):
            vehicle_message_name = vehicle_message_name_column[i+1].value
            vehicle_signal_name = vehicle_signal_name_column[i+1].value
            ecu_message_name = ecu_message_name_column[i+1].value
            ecu_signal_name = ecu_signal_name_column[i+1].value
            if vehicle_message_name != None and vehicle_signal_name != None and ecu_message_name != None and ecu_signal_name != None:
                vehicle_message_name = vehicle_message_name.strip()
                vehicle_signal_name = vehicle_signal_name.strip()
                ecu_message_name = ecu_message_name.strip()
                ecu_signal_name = ecu_signal_name.strip()
            sheet.cell(i+2, 12).value = f'{vehicle_message_name}.{vehicle_signal_name}.phys = '
            sheet.cell(i+2, 13).value = '' if ecu_signal_name == None else f'{ecu_message_name}.{ecu_signal_name} = '
            c_code = ""
            output_code = []
            cell_value = cell.value
            if not cell_value:
                continue
            elif cell_value.startswith("Default"):
                c_code = cell_value
            elif cell_value.startswith("Direct"):
                c_code = ""
                try:
                    nums = re.findall(r'0x[\da-fA-F]+', sheet["B"][i + 1].value)
                    for a in nums:
                        if a != 0:
                            input_value = int(a, 16)
                            sheet.cell(i+2, 12).value = f'{vehicle_message_name}.{vehicle_signal_name}.phys = {input_value}'
                            sheet.cell(i+2, 13).value = f'{ecu_message_name}.{ecu_signal_name} = {input_value}'
                except Exception as e:
                    pass            
            elif cell_value.startswith("Case"):
                case_statements = cell_value.split(";")
                Default_exist = False
                for case_statement in case_statements:
                    if case_statement.find('；') != -1:
                        logging.error(f'{sheet}第{i+2}行包含中文分号')
                    elif case_statement.find('：') != -1:
                        logging.error(f'{sheet}第{i+2}行包含中文冒号')
                    elif "Case" in case_statement:
                        hex_num = re.findall(r'0x[\da-fA-F]+', case_statement)
                        condition = int(hex_num[0], 16)
                        action = int(hex_num[1], 16)
                        if condition != 0:
                            sheet.cell(i+2, 12).value = f'{vehicle_message_name}.{vehicle_signal_name}.phys = {condition}'
                            sheet.cell(i+2, 13).value = f'{ecu_message_name}.{ecu_signal_name} = {action}'
                        output_code.append(f"        case {condition}:")
                        output_code.append(f"            l_temp = {action};")
                        output_code.append("            break;")
                    elif case_statement.startswith("\nDefault"):
                        action = int(re.findall(r'0x[\da-fA-F]+', case_statement)[0], 16)
                        output_code.append("        default:")
                        output_code.append(f"            l_temp = {action};")
                        output_code.append("            break;")
                        Default_exist = True
                    elif case_statement == "":
                        pass
                    else:
                        logging.error(f'{sheet}第{i+2}行缺失"Case"')
                if not Default_exist:
                    logging.error(f'{sheet}第{i+2}行缺失Default')
                temp = '\n'.join(output_code)
                c_code = f'    switch(l_temp)\n    {{\n{temp}\n    }}'
            elif cell_value.startswith("MAX"):
                case_statements = cell_value.split(";")
                for case_statement in case_statements:
                    if case_statement.find('；') != -1:
                        logging.error(f'{sheet}第{i+2}行包含中文分号')
                    if case_statement.find('：') != -1:
                        logging.error(f'{sheet}第{i+2}行包含中文冒号')
                    elif "MAX" in case_statement:
                        max_value = case_statement.split(":")[1].strip()
                        output_code.append(f'    if(l_temp >= {max_value})\n        l_temp = {max_value};\n')
                    elif case_statement.startswith("\nMIN"):
                        min_value = case_statement.split(":")[1].strip()
                        output_code.append(f'    else if(l_temp <= {min_value})\n       l_temp = {min_value};\n')
                    elif case_statement.startswith("\nvalue"):
                        coefficient = case_statement.replace("value", "").strip()
                        output_code.append(f'    l_temp = l_temp{coefficient};')
                    elif case_statement == "":
                        pass
                    else:
                        logging.error(f'{sheet}第{i+2}行的Mapping Rule无法识别')
                c_code = '\n'.join(output_code)
            elif cell_value.startswith("value"):
                coefficient = cell_value.replace("value", "").strip()
                c_code = f'    l_temp = l_temp{coefficient};\n'
            elif cell_value.startswith("signal"):
                case_statements = cell_value.split(";")
                for case_statement in case_statements:
                    if case_statement.find('；') != -1:
                            logging.error(f"{sheet}第{i+2}行包含中文分号")
                    elif case_statement.find('：') != -1:
                        logging.error(f'{sheet}第{i+2}行包含中文冒号')
                    elif "signal1" in case_statement:
                        if "MAX" in case_statement:
                            max_value = case_statement.split(":")[2].strip()
                            output_code.append(f'    if(l_temp >= {max_value})\n        l_temp = {max_value};\n')
                        elif "MIN" in case_statement:
                            min_value = case_statement.split(":")[2].strip()
                            output_code.append(f'    else if(l_temp <= {min_value})\n       l_temp = {min_value};\n')
                        elif "value" in case_statement:
                            condition = case_statement.split(":")[1].replace("value=", "").strip()
                            action = case_statement.split(":")[1].replace("value", "").strip()
                            output_code.append(f'    if(l_temp == {condition})\n    {{\n        l_temp = l_temp{action};\n    }}')
                        elif "Direct" in case_statement:
                            c_code = ""
                        elif "Case" in case_statement:
                            output_code.append("// TODO")
                    elif "signal2" in case_statement:
                        m += 1
                        vehicle_message_name_1 = vehicle_message_name_column[i+1+m].value
                        vehicle_signal_name_1 = vehicle_signal_name_column[i+1+m].value
                        if vehicle_message_name_1 != None and vehicle_signal_name_1 != None:
                            vehicle_message_name_1 = vehicle_message_name_1.strip()
                            vehicle_signal_name_1 = vehicle_signal_name_1.strip()
                        signal_info = f"{device}_{camel_to_underscore(vehicle_message_name_1).lower()}_{camel_to_underscore(vehicle_signal_name_1)}_decode(g_{device}_{vehicle_message_name_1.lower()}.{camel_to_underscore(vehicle_signal_name_1)})"
                        if "MAX" in case_statement:
                            max_value = case_statement.split(":")[2].strip()
                            output_code.append(f'    if({signal_info} >= {max_value})\n        l_temp = {max_value};\n')
                        elif "MIN" in case_statement:
                            min_value = case_statement.split(":")[2].strip()
                            output_code.append(f'    else if({signal_info} <= {min_value})\n       l_temp = {min_value};\n')
                        elif "value" in case_statement:
                            condition = case_statement.split(":")[1].replace("value=", "").strip()
                            action = case_statement.split(":")[2].replace("value", "").strip()
                            output_code.append(f'    if({signal_info} == {condition})\n    {{\n        l_temp = l_temp{action};\n    }}')
                        elif cell_value.startswith("Direct"):
                            c_code = ""
                        elif "Case" in case_statement:
                            output_code.append("// TODO")         
                        elif case_statement == "":
                                pass
                        else:
                            logging.error(f'{sheet}第{i+2}行的Mapping Rule无法识别')
                    elif case_statement == "":
                        pass
                    else:
                        logging.error(f'{sheet}第{i+2}行的Mapping Rule无法识别')
                c_code = '\n'.join(output_code)
            else:
                logging.error(f'{sheet}第{i+2}行的Mapping Rule无法识别')
                c_code = "// TODO"
            print (c_code)
            if vehicle_signal_name != None and vehicle_message_name!= None:
                information = [f"{device}_{camel_to_underscore(vehicle_message_name).lower()}_{camel_to_underscore(vehicle_signal_name)}_decode(g_{device}_{vehicle_message_name.lower()}.{camel_to_underscore(vehicle_signal_name)})", f'{ecu_message_name}_{ecu_signal_name}', c_code]        
            else:
                information = []
            infors.append(information)
        
        if not os.path.exists('./output'):
            os.makedirs('./output')
        self.workbook.save('./output/Gateway_template.xlsx')
        return infors
        
def main(source_folder_path, mapping_rule_excel):
    if not os.path.exists('./output'):
        os.makedirs('./output')
    logging.basicConfig(filename='./output/mapping_rule_error.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding = 'utf-8') # , encoding = 'utf-8'
    
    excel_check(source_folder_path, mapping_rule_excel)
    parser = ExcelParser(mapping_rule_excel)
    signal_length_dict = parse_dbc(source_folder_path)[0]
    generated_code = parser.parse_mapping_rules(2, signal_length_dict, 'ecu')
    generated_code = parser.parse_mapping_rules(3, signal_length_dict, 'vehicle')

    logging.shutdown()

if __name__ == "__main__":
    main(source_folder_path='.\input\_DBC', mapping_rule_excel='.\Gateway_template.xlsx')