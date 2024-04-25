import openpyxl
import re

def first_check(sheet, pattern, hex_pattern, separator, log_file):
    log_file.write(separator + '\n' + "Mapping excel error about empty None:" + '\n')
    for column_index in (0, 2, 3, 5, 7, 8, 9):  
        column_letter = openpyxl.utils.get_column_letter(column_index + 1)
        for cell in sheet[column_letter][1:]:
            cell_value = cell.value
            if cell_value is None:
                log_message = f"Cell in Row {cell.row} and Column {column_letter} is empty (None)."
                log_file.write(log_message + '\n')

    log_file.write(separator +'\n'+"Mapping excel error about Carriage returns and spaces :" +'\n')
    pattern = re.compile(r'[ \r\n\s]+')
    for column_index in (0, 2, 3, 5, 7):  
        column_letter = openpyxl.utils.get_column_letter(column_index + 1)
        for cell in sheet[column_letter][1:]:
            cell_value = cell.value
            if cell_value and isinstance(cell_value, str):
                match = pattern.search(cell_value)
                if match:

                    cleaned_value = pattern.sub('', cell_value)
                    cell.value = cleaned_value  
                    # log_file.write(separator + '\n')
                    log_message = f"Cell Value: '{cell_value}' in Row {cell.row} and Column {column_letter} had extra whitespace or newline characters and was cleaned."
                    log_file.write(log_message + '\n')


    log_file.write(separator+'\n' + 'Mapping excel error about hexadecimal :' +'\n')
    for column_index in (3, 8):  # 对应D列和I列的索引
        column_letter = openpyxl.utils.get_column_letter(column_index + 1)
        for cell in sheet[column_letter][1:]:
            cell_value = cell.value
            if isinstance(cell_value, int):
                cell_value = str(cell_value)
            if cell_value and isinstance(cell_value, str):
                if cell_value.startswith("0x") or cell_value.startswith("0X"):
                    # 如果是以 "0x" 或 "0X" 开头，继续检查是否是有效的十六进制
                    if not hex_pattern.match(cell_value):
                        log_message = f"Cell Value: '{cell_value}' in Row {cell.row} and Column {column_letter} is not in a valid hexadecimal format."
                        log_file.write(log_message + '\n')
                else:
                    log_message = f"Cell Value: '{cell_value}' in Row {cell.row} and Column {column_letter} is not prefixed with '0x' or '0X' and is not a valid hexadecimal format."
                    cell.value = "0x" + cell_value
                    log_file.write(log_message + '\n')

    log_file.write(separator+'\n' + 'Mapping excel error about Duplicate values :' +'\n')                    
    columns_to_check = [1, 6] 
    column_values = {col: [] for col in columns_to_check}
    for row in sheet.iter_rows(values_only=True):
        for col_index, value in enumerate(row):
            if col_index + 1 in columns_to_check:
                if isinstance(value, str):
                    column_values[col_index + 1].append(value)
    for col_index in columns_to_check:
        col_letter = openpyxl.utils.get_column_letter(col_index)
        duplicates = [item for item in column_values[col_index] if column_values[col_index].count(item) > 1]
        if duplicates:
            log_message = f"Duplicate values in {col_letter} column (Column {col_index}): {', '.join(map(str, set(duplicates)))}"
            log_file.write(log_message + '\n')

    return

def main(excel_file):
    workbook = openpyxl.load_workbook(excel_file)  
    sheet_to_check = workbook.worksheets
    pattern = re.compile(r'[ \r\n\s]+')
    hex_pattern = re.compile(r'^0[xX]([0-9A-Fa-f]+)$')
    separator = '-' * 400
    log_file = open('./output/First check.txt', 'w', encoding='utf-8')
    for sheet in [sheet_to_check[2],sheet_to_check[3]]:
        log_file.write(f"Checking sheet: {sheet.title}\n" + separator + '\n')
        first_check(sheet, pattern, hex_pattern, separator, log_file)
    workbook.save('./output/cleaned_excel_file.xlsx')
    print('First check has been finished')
    log_file.close()

# Check all error about:空缺值，空格回车符，十六进制，重复值
# Input: ./Gateway_templete.xlsx
# Output: log.txt

if __name__ == "__main__":
    main('./Gateway_template.xlsx')
